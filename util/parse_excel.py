#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2019/10/14 23:37
# @Author  : Tang Yiwei
# @Email   : 892398433@qq.com
# @File    : parse_excel.py
# @Software: PyCharm

import openpyxl
from openpyxl.styles import Border,Side,Alignment,Font
import time

class ParseExcel():
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.worksheet = None #add
        self.fontColor = Font(color = None) # 设置字体的颜色
        self.fontStyle = Font(name = None) # 设置字体的样式
        # 设置字体样式
        self.fontStyleDict = {
            "microsoft_accor_black":"微软雅黑",
            "regular_script":"楷体",
            "song_typeface":"宋体"
        }
        # 颜色对应的RGB值
        self.RGBDict = {"red":"FFFF3030",
                        "green":"FF008B00",
                        "black":"000000",
                        }

    def createExcel(self,excelPathName):
        # 新建excel
        try:
            self.workbook = openpyxl.Workbook()
        except Exception as e:
            raise e
        self.excelFile = excelPathName
        return self.workbook

    def active(self):
        #  获取活跃的表格
        self.worksheet = self.workbook.active
        return self.worksheet

    def addTitle(self,title):
        # 修改title
        self.active.title = title

    def addSheet(self,sheetName):
        # 增加sheet
        self.worksheet_n = self.workbook.create_sheet(sheetName)

    def loadWordBook(self,excelPathName):
        # 将excel加载到内存中，并获取其workbook对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathName)
        except Exception as e:
            raise e
        self.excelFile = excelPathName
        return self.workbook

    def getSheetByName(self,sheetName):
        # 根据sheet名获取该sheet对象
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self,sheetIndex):
        # 根据sheet的索引号获取该sheet对象
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self,sheet):
        # 获取sheet中有数据区域的结束行号
        return sheet.max_row

    def getColsNumber(self,sheet):
        # 获取sheet中有数据区域的结束列号
        return sheet.max_column

    def getStartRowNumber(self,sheet):
        # 获取sheet中有数据区域的开始行号
        return sheet.min_row

    def getStartColNumber(self,sheet):
        # 获取sheet中有数据区域开始的列号
        return sheet.min_column

    def detele_rows(self,sheet,startRow,N):
        # 删除行，删除startRow开始以后的N行
        return sheet.delete_rows(startRow,N)

    def getRow(self,sheet,rowNo):
        # 获取sheet中某一行，返回的是这一行所有的数据内容组成tuple
        # 下标从1开始，sheet,row[1]表示第一行
        dataList = []
        rows = self.getRowsNumber(sheet)
        cols = self.getColsNumber(sheet)
        if rowNo - 1 >= rows:
            print('row out of max range! Please Check excel row number!')
        try:
            for line in sheet.rows:
                tmpList = []
                for i in range(cols):
                    tmpList.append(line[i].value)
                dataList.append(tmpList)
            return dataList[rowNo - 1]
        except Exception as e:
            raise e

    def getColumn(self,sheet,colNo):
        # 获取sheet中某一列，返回的是这一列所有的数据内容组成tuple
        # 下标从1开始，sheet,row[1]表示第一列
        dataList = []
        rows = self.getRowsNumber(sheet)
        cols = self.getColsNumber(sheet)
        if colNo - 1 >= cols:
            print('column out of max range! Please Check excel column number!')
        try:
            for line in sheet.columns:
                tmpList = []
                for i in range(1,rows):#过滤掉第一列的标题数据
                    tmpList.append(line[i].value)
                dataList.append(tmpList)
            return tuple(dataList[colNo - 1])
        except Exception as e:
            raise e

    def getCellOfValue(self,sheet,coordinate = None,rowNo = None,colsNo = None):
        # 根据单元格所在位置索引获取该单元格中的值，下标从1开始
        # sheet.cell(row  = 1,column = 1).value,表示Excel中第一行第一列的值
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate).value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo,column = colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell")

    def getCellOfObject(self,sheet,coordinate = None,rowNo = None,colsNo = None):
        # 获取某个单元格对的对象，可以根据单元格所在位置的数字索引
        # 也可以直接根据excel中单元格的编码及坐标
        # 如getCellOfObject(sheet,coordinate = 'A1',) or
        # getCellOfObject(sheet,rowNo = 1,colsNo = 2)
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate)
            except Exception as e:
                raise e

        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo,column = colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception('Insufficient Coordinates of cell')

    def writeCell(self,sheet,content,coordinate = None,rowNo = None,colsNo = None,colorStyle=None,fontStyle=None):
        # 根据单元格在Excel中的编码坐标或者数字索引坐标向单元格中写入数据
        # 下标从1开始，参数style表示字体的颜色的名字，比如red,green
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = content
                if colorStyle is not None:
                    sheet.cell(coordinate = coordinate).font = Font(color = self.RGBDict[colorStyle])
                if fontStyle is not None:
                    sheet.cell(coordinate = coordinate).font = Font(name = self.fontStyleDict[fontStyle])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo,column = colsNo).value = content
                if colorStyle:
                    sheet.cell(row = rowNo,column = colsNo).font = Font(color = self.RGBDict[colorStyle])
                if fontStyle:
                    sheet.cell(coordinate=coordinate).font = Font(name=self.fontStyleDict[fontStyle])
                self.workbook.save(self.excelFile)
            except PermissionError as e:
                print("file is open or Please check file permissions!")
            except Exception as e:
                raise Exception('Insufficient Coordinates of cell')
        else:
            raise Exception('Insufficient Coordinates of cell')

    def writeCellCurrentTime(self,sheet,coordinate = None,rowNo = None,colsNo = None):
        # 写入当前时间，下标从1开始
        now = int(time.time()) #显示未时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo,column = colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except PermissionError as e:
                print("file is open or Please check file permissions!")
            except Exception as e:
                raise e
        else:
            raise Exception('Insufficient Coordinates of cell')

    def writeCelltmp(self, sheet, content, coordinate=None, rowNo=None, colsNo=None, colorStyle=None,fontStyle=None):
        # 只写入数据，不保存数据，保存数据统一调用saveTable()方法
        # 根据单元格在Excel中的编码坐标或者数字索引坐标向单元格中写入数据
        # 下标从1开始，参数style表示字体的颜色的名字，比如red,green
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if colorStyle is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGBDict[colorStyle])
                if fontStyle is not None:
                    sheet.cell(coordinate=coordinate).font = Font(name=self.fontStyleDict[fontStyle])
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = content
                if colorStyle:
                    sheet.cell(row=rowNo, column=colsNo).font = Font(color=self.RGBDict[colorStyle])
                if fontStyle:
                    sheet.cell(row=rowNo, column=colsNo).font = Font(name=self.fontStyleDict[fontStyle])
            except PermissionError as e:
                print("file is open or Please check file permissions!")
            except Exception as e:
                raise Exception('Insufficient Coordinates of cell')
        else:
            raise Exception('Insufficient Coordinates of cell')

    def saveTable(self):
        # 保存表格
        self.workbook.save(self.excelFile)



if __name__ == "__main__":
    pass